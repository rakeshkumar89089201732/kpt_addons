# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime, timedelta, time
import pytz

try:
    import openpyxl
except ImportError:
    openpyxl = None

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

class BulkAttendanceWizard(models.TransientModel):
    _name = 'bulk.attendance.wizard'
    _description = 'Bulk Attendance Upload'

    upload_file = fields.Binary(string='Upload File', required=True)
    file_name = fields.Char(string='File Name')
    mode = fields.Selection([
        ('day', 'Daily Logs'),
        ('month', 'Monthly Muster Roll')
    ], string='Upload Mode', default='day', required=True)
    
    attendance_date = fields.Date(
        string='Date', 
        default=fields.Date.context_today,
        help="For 'Daily Logs', this date is used if dates are missing in file. "
             "For 'Monthly Muster Roll', this determines the Month/Year."
    )
    
    default_check_in = fields.Float(string='Default Check In', default=9.0, help="Time in hours (e.g. 9.0 = 09:00 AM)")
    default_check_out = fields.Float(string='Default Check Out', default=18.0, help="Time in hours (e.g. 18.0 = 06:00 PM)")

    def action_import(self):
        """Import attendance"""
        if not openpyxl:
            raise UserError(_("The 'openpyxl' library is required to import Excel files."))
            
        if not self.upload_file:
            raise UserError(_("Please upload a file."))

        try:
            file_data = base64.b64decode(self.upload_file)
            file_input = io.BytesIO(file_data)
            workbook = openpyxl.load_workbook(file_input, data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid file format. Please upload a valid Excel file (.xlsx). Error: %s") % str(e))

        if self.mode == 'day':
            self._import_day(sheet)
        else:
            self._import_month(sheet)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Success'),
                'message': _('Attendance records imported successfully.'),
                'type': 'success',
                'sticky': False,
            }
        }

    def _get_allowed_employee_ids(self):
        """Get list of employee IDs the current user is allowed to upload attendance for.
        For managers with team access, returns their subordinates. For others, returns all employees.
        """
        if hasattr(self.env.user, '_get_restricted_employee_ids'):
            allowed_ids = self.env.user._get_restricted_employee_ids()
            if allowed_ids is not None:
                return allowed_ids
        # If no restriction or method doesn't exist, return None (allow all)
        return None

    def _is_employee_allowed(self, employee):
        """Check if the current user is allowed to upload attendance for this employee."""
        allowed_ids = self._get_allowed_employee_ids()
        if allowed_ids is None:
            return True  # No restriction, allow all
        return employee.id in allowed_ids

    def _get_employee(self, code):
        """Find employee by code (barcode, pin, or name)"""
        if not code:
            return None
        code = str(code).strip()
        domain = ['|', '|', ('barcode', '=', code), ('pin', '=', code), ('name', '=', code)]
        # Add internal reference if exists
        # if hasattr(self.env['hr.employee'], 'registration_number'):
        #      domain = ['|'] + domain + [('registration_number', '=', code)]
             
        employee = self.env['hr.employee'].search(domain, limit=1)
        return employee

    def _float_to_time(self, float_val):
        """Convert float hour to time object"""
        hours = int(float_val)
        minutes = int((float_val - hours) * 60)
        return time(hours, minutes, 0)
    
    def _localize_dt(self, dt):
        """Convert naive datetime to UTC based on user's timezone"""
        user_tz = self.env.user.tz or 'UTC'
        local = pytz.timezone(user_tz)
        local_dt = local.localize(dt, is_dst=None)
        return local_dt.astimezone(pytz.UTC).replace(tzinfo=None)

    def _import_day(self, sheet):
        """
        Expected Columns: Employee, Check In, Check Out
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
             raise UserError(_("The file is empty."))
             
        header = [str(x).lower().strip() if x else '' for x in rows[0]]
        
        col_emp = -1
        col_in = -1
        col_out = -1
        
        for i, h in enumerate(header):
            if 'employee' in h or 'code' in h:
                col_emp = i
            elif 'in' in h:
                col_in = i
            elif 'out' in h:
                col_out = i

        if col_emp == -1 or col_in == -1:
            raise UserError(_("Missing required columns: Employee Code, Check In"))

        count = 0
        skipped_unauthorized = []
        for row_idx, row in enumerate(rows[1:], start=2):
            emp_code = row[col_emp]
            employee = self._get_employee(emp_code)
            
            if not employee:
                _logger.info("Row %s: Employee not found for code %s", row_idx, emp_code)
                continue

            # Check if user is allowed to upload attendance for this employee
            if not self._is_employee_allowed(employee):
                skipped_unauthorized.append(emp_code)
                _logger.info("Row %s: User not authorized to upload attendance for employee %s", row_idx, emp_code)
                continue

            check_in_val = row[col_in]
            check_out_val = row[col_out] if col_out != -1 else None

            # Parse Check In
            if isinstance(check_in_val, datetime):
                dt_in = check_in_val
            elif isinstance(check_in_val, time):
                dt_in = datetime.combine(self.attendance_date, check_in_val)
            else:
                 _logger.info("Row %s: Invalid Check In format", row_idx)
                 continue

            # Parse Check Out
            dt_out = False
            if check_out_val:
                if isinstance(check_out_val, datetime):
                    dt_out = check_out_val
                elif isinstance(check_out_val, time):
                    dt_out = datetime.combine(self.attendance_date, check_out_val)

            # Convert to UTC
            dt_in = self._localize_dt(dt_in)
            if dt_out:
                dt_out = self._localize_dt(dt_out)

            self.env['hr.attendance'].create({
                'employee_id': employee.id,
                'check_in': dt_in,
                'check_out': dt_out,
            })
            count += 1
        
        if skipped_unauthorized:
            _logger.warning("Skipped %d unauthorized employee(s): %s", len(skipped_unauthorized), ', '.join(set(skipped_unauthorized)))
            
        if count == 0:
            if skipped_unauthorized:
                raise UserError(_("No valid records imported. You are not authorized to upload attendance for the employees in this file. Please contact your administrator."))
            raise UserError(_("No valid records imported."))

    def _import_month(self, sheet):
        """
        Expected Columns: Employee Code, 1, 2, 3, ... (Days)
        Values: 'P' or 'Present'
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
             raise UserError(_("The file is empty."))

        header = [str(x).lower().strip() if x else '' for x in rows[0]]
        
        col_emp = -1
        day_cols = {}
        
        for i, h in enumerate(header):
            if 'employee' in h or 'code' in h:
                col_emp = i
                continue
            
            try:
                day = int(h)
                if 1 <= day <= 31:
                    day_cols[day] = i
            except ValueError:
                pass
                
        if col_emp == -1:
             raise UserError(_("Missing Employee Code column."))
        if not day_cols:
             raise UserError(_("No day columns (1-31) found."))

        base_date = self.attendance_date
        year = base_date.year
        month = base_date.month
        
        t_in = self._float_to_time(self.default_check_in)
        t_out = self._float_to_time(self.default_check_out)

        count = 0
        skipped_unauthorized = []
        for row_idx, row in enumerate(rows[1:], start=2):
            emp_code = row[col_emp]
            employee = self._get_employee(emp_code)
            
            if not employee:
                continue

            # Check if user is allowed to upload attendance for this employee
            if not self._is_employee_allowed(employee):
                skipped_unauthorized.append(emp_code)
                _logger.info("Row %s: User not authorized to upload attendance for employee %s", row_idx, emp_code)
                continue

            for day, col_idx in day_cols.items():
                if col_idx >= len(row): 
                    continue
                    
                val = str(row[col_idx]).strip().upper() if row[col_idx] else ''
                
                # Check for Present markers
                if val in ['P', 'PRESENT', 'YES', '1', '1.0']:
                    try:
                        current_date = datetime(year, month, day).date()
                    except ValueError:
                        continue # Invalid date

                    # Determine times
                    # TODO: Resource calendar logic could go here
                    dt_in = datetime.combine(current_date, t_in)
                    dt_out = datetime.combine(current_date, t_out)
                    
                    # Deduplicate: Check if attendance exists for this day
                    # We need to check in UTC.
                    
                    # Convert to UTC for storage
                    utc_in = self._localize_dt(dt_in)
                    utc_out = self._localize_dt(dt_out)
                    
                    # Search range (full day in UTC) needs care, but simple existence check is safer
                    # Search for any attendance that overlaps or starts on this day?
                    # Simplest: Check if any attendance starts on this day.
                    
                    # Convert user local start/end of day to UTC to search
                    user_tz = pytz.timezone(self.env.user.tz or 'UTC')
                    local_day_start = user_tz.localize(datetime.combine(current_date, time.min))
                    local_day_end = user_tz.localize(datetime.combine(current_date, time.max))
                    
                    utc_start = local_day_start.astimezone(pytz.UTC).replace(tzinfo=None)
                    utc_end = local_day_end.astimezone(pytz.UTC).replace(tzinfo=None)

                    exists = self.env['hr.attendance'].search_count([
                        ('employee_id', '=', employee.id),
                        ('check_in', '>=', utc_start),
                        ('check_in', '<=', utc_end),
                    ])
                    
                    if not exists:
                        self.env['hr.attendance'].create({
                            'employee_id': employee.id,
                            'check_in': utc_in,
                            'check_out': utc_out,
                        })
                        count += 1
        
        if skipped_unauthorized:
            _logger.warning("Skipped %d unauthorized employee(s): %s", len(skipped_unauthorized), ', '.join(set(skipped_unauthorized)))
        
        if count == 0:
            if skipped_unauthorized:
                raise UserError(_("No valid records imported. You are not authorized to upload attendance for the employees in this file. Please contact your administrator."))
            raise UserError(_("No valid records imported."))
