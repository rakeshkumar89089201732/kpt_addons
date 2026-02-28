# -*- coding: utf-8 -*-

import base64
import io
import pytz
from calendar import monthrange
from datetime import date, datetime, time, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None

from odoo import _, api, fields, models
from odoo.exceptions import UserError


class MonthlyAttendanceConfigWizard(models.TransientModel):
    """Single Wizard for Monthly Attendance Creation"""
    _name = 'monthly.attendance.config.wizard'
    _description = 'Monthly Attendance Configuration Wizard'

    year = fields.Char(
        string='Year',
        required=True,
        default=lambda self: str(datetime.now().year),
    )
    month = fields.Selection(
        [
            ('01', 'January'), ('02', 'February'), ('03', 'March'),
            ('04', 'April'), ('05', 'May'), ('06', 'June'),
            ('07', 'July'), ('08', 'August'), ('09', 'September'),
            ('10', 'October'), ('11', 'November'), ('12', 'December'),
        ],
        string='Month',
        required=True,
        default=lambda self: str(datetime.now().month).zfill(2),
    )
    resource_calendar_id = fields.Many2one(
        'resource.calendar',
        string='Working Schedule',
        required=True,
        help='Working schedule to use for calculating default hours and holidays. Auto-selected from company default.',
    )
    excel_file = fields.Binary(string='Excel File')
    excel_filename = fields.Char(string='Excel Filename')
    default_check_in = fields.Float(
        string='Default Check In (Hours)',
        compute='_compute_default_hours',
        readonly=True,
        help='Default check-in time from the selected working schedule.',
    )
    default_check_out = fields.Float(
        string='Default Check Out (Hours)',
        compute='_compute_default_hours',
        readonly=True,
        help='Default check-out time from the selected working schedule.',
    )
    default_hours_per_day = fields.Float(
        string='Default Hours per Day',
        compute='_compute_default_hours',
        readonly=True,
        help='Average working hours per day from the selected schedule.',
    )
    total_sundays = fields.Integer(
        string='Total Sundays',
        compute='_compute_holidays',
        readonly=True,
    )
    total_holidays = fields.Integer(
        string='Total Public Holidays',
        compute='_compute_holidays',
        readonly=True,
    )
    total_working_days = fields.Integer(
        string='Total Working Days',
        compute='_compute_holidays',
        readonly=True,
    )
    
    employee_line_ids = fields.One2many(
        'monthly.attendance.employee.line',
        'wizard_id',
        string='Employees',
    )

    @api.model
    def default_get(self, fields_list):
        """Auto-select company's default working schedule"""
        res = super().default_get(fields_list)
        if 'resource_calendar_id' in fields_list and not res.get('resource_calendar_id'):
            # Get company's default calendar
            company = self.env.company
            if company.resource_calendar_id:
                res['resource_calendar_id'] = company.resource_calendar_id.id
            else:
                # Fallback to first available calendar
                calendar = self.env['resource.calendar'].search([], limit=1)
                if calendar:
                    res['resource_calendar_id'] = calendar.id
        return res

    @api.depends('resource_calendar_id')
    def _compute_default_hours(self):
        """Calculate default hours from the selected working schedule"""
        for record in self:
            if not record.resource_calendar_id:
                record.default_check_in = 9.0
                record.default_check_out = 18.0
                record.default_hours_per_day = 8.0
                continue

            calendar = record.resource_calendar_id
            # Get average hours per day
            hours_per_day = calendar.hours_per_day
            
            # Get first working day's schedule to determine check-in/check-out
            attendances = calendar.attendance_ids.filtered(lambda a: a.dayofweek == '0')  # Monday
            if not attendances:
                attendances = calendar.attendance_ids[:1]
            
            if attendances:
                # Get earliest start and latest end
                start_times = [a.hour_from for a in attendances]
                end_times = [a.hour_to for a in attendances]
                if start_times and end_times:
                    record.default_check_in = min(start_times)
                    record.default_check_out = max(end_times)
                else:
                    record.default_check_in = 9.0
                    record.default_check_out = 18.0
            else:
                record.default_check_in = 9.0
                record.default_check_out = 18.0
            
            record.default_hours_per_day = hours_per_day or 8.0

    @api.depends('year', 'month', 'resource_calendar_id')
    def _compute_holidays(self):
        """Calculate total Sundays and public holidays for the selected month"""
        for record in self:
            if not record.year or not record.month:
                record.total_sundays = 0
                record.total_holidays = 0
                record.total_working_days = 0
                continue

            try:
                y = int(record.year)
                m = int(record.month)
            except (ValueError, TypeError):
                record.total_sundays = 0
                record.total_holidays = 0
                record.total_working_days = 0
                continue

            last_day = monthrange(y, m)[1]
            date_from = date(y, m, 1)
            date_to = date(y, m, last_day)

            # Count Sundays
            sundays = 0
            d = date_from
            while d <= date_to:
                if d.weekday() == 6:  # Sunday = 6
                    sundays += 1
                d += timedelta(days=1)
            record.total_sundays = sundays

            # Count public holidays from calendar
            holidays = 0
            holiday_dates = set()
            if record.resource_calendar_id:
                leaves = self.env['resource.calendar.leaves'].search([
                    ('calendar_id', '=', record.resource_calendar_id.id),
                    ('date_from', '<=', datetime.combine(date_to, datetime.max.time())),
                    ('date_to', '>=', datetime.combine(date_from, datetime.min.time())),
                ])
                for leave in leaves:
                    start_dt = fields.Datetime.to_datetime(leave.date_from)
                    end_dt = fields.Datetime.to_datetime(leave.date_to)
                    if start_dt and end_dt:
                        cur = start_dt.date()
                        end = end_dt.date()
                        while cur <= end:
                            if date_from <= cur <= date_to:
                                holiday_dates.add(cur)
                            cur += timedelta(days=1)
                holidays = len(holiday_dates)
            record.total_holidays = holidays

            # Calculate working days (excluding Sundays and holidays)
            working_days = 0
            d = date_from
            while d <= date_to:
                if d.weekday() != 6 and d not in holiday_dates:  # Not Sunday and not holiday
                    working_days += 1
                d += timedelta(days=1)
            record.total_working_days = working_days

    def action_create_attendance(self):
        """Create attendance records for all employees"""
        self.ensure_one()
        if not self.resource_calendar_id:
            raise UserError(_('Please select a working schedule.'))
        
        if not self.employee_line_ids:
            raise UserError(_('Please add at least one employee.'))

        created_count = 0
        errors = []

        for line in self.employee_line_ids:
            if line.total_present_days <= 0:
                errors.append(_('Employee %s: Total Present Days must be greater than 0.') % line.employee_id.name)
                continue

            try:
                count = line.create_attendance_records()
                created_count += count
            except Exception as e:
                errors.append(_('Employee %s: %s') % (line.employee_id.name, str(e)))

        message = _('Created attendance records for %d employee(s).') % created_count
        if errors:
            message += '\n\n' + _('Errors:') + '\n' + '\n'.join(errors)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Success') if not errors else _('Completed with Errors'),
                'message': message,
                'type': 'success' if not errors else 'warning',
                'sticky': True,
            },
        }

    def action_cancel_and_back(self):
        """Cancel wizard and go back to where it was opened from.

        - If opened from dashboard (context flag), go back to /attendance/dashboard
        - Otherwise, behave like a normal wizard (close and return to previous action)
        """
        if self.env.context.get('from_attendance_dashboard'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/attendance/dashboard',
                'target': 'self',
            }
        # default wizard close behaviour
        return {'type': 'ir.actions.act_window_close'}

    def _get_employee_domain(self):
        """Get domain for employee field based on manager restrictions."""
        domain = [('active', '=', True)]
        if hasattr(self.env.user, '_get_restricted_employee_ids'):
            allowed_ids = self.env.user._get_restricted_employee_ids()
            if allowed_ids is not None:
                if not allowed_ids:
                    domain.append(('id', '=', False))  # No access
                else:
                    domain.append(('id', 'in', allowed_ids))
        return domain

    def action_export_template(self):
        """Export Excel template with employees accessible to the current user (manager's subordinates)"""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("The 'openpyxl' library is required to export Excel files."))

        # Get employees based on manager restrictions (only manager's subordinates)
        domain = self._get_employee_domain()
        employees = self.env['hr.employee'].search(domain, order='name')
        
        if not employees:
            raise UserError(_("No employees available for export. You may not have access to any employees or there are no active employees in your team."))
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Attendance"
        
        # Header row
        headers = ['Employee Name', 'Employee Code', 'Total Present Days', 'Total Overtime (Hours)', 
                   'Check In (Hours)', 'Check Out (Hours)']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        row_idx = 2
        for employee in employees:
            ws.cell(row=row_idx, column=1, value=employee.name)
            ws.cell(row=row_idx, column=2, value=employee.barcode or employee.pin or '')
            ws.cell(row=row_idx, column=3, value=0)  # Total Present Days
            ws.cell(row=row_idx, column=4, value=0.0)  # Total Overtime
            ws.cell(row=row_idx, column=5, value=self.default_check_in or 9.0)  # Check In
            ws.cell(row=row_idx, column=6, value=self.default_check_out or 18.0)  # Check Out
            row_idx += 1
        
        # Auto-adjust column widths
        column_widths = [30, 15, 18, 22, 18, 18]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        month_name = dict(self._fields['month'].selection).get(self.month, self.month)
        filename = f"Monthly_Attendance_Template_{month_name}_{self.year}.xlsx"
        
        # Encode to base64
        file_data = base64.b64encode(output.read())
        
        # Create attachment for download
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def action_import_excel(self):
        """Import employee data from Excel file"""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("The 'openpyxl' library is required to import Excel files."))
        
        if not self.excel_file:
            raise UserError(_("Please upload an Excel file."))
        
        try:
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid Excel file. Error: %s") % str(e))
        
        # Clear existing lines
        self.employee_line_ids.unlink()
        
        # Read data (skip header row)
        lines_to_create = []
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            employee_name = str(row[0]).strip() if row[0] else None
            employee_code = str(row[1]).strip() if len(row) > 1 and row[1] else None
            
            if not employee_name and not employee_code:
                continue
            
            # Find employee - try multiple methods (with manager restrictions)
            employee = None
            base_domain = self._get_employee_domain()
            
            if employee_code:
                # Try by code first (barcode or pin)
                search_domain = base_domain + ['|', ('barcode', '=', employee_code), ('pin', '=', employee_code)]
                employee = self.env['hr.employee'].search(search_domain, limit=1)
            
            # If not found by code, try by name
            if not employee and employee_name:
                search_domain = base_domain + [('name', '=', employee_name)]
                employee = self.env['hr.employee'].search(search_domain, limit=1)
            
            # If still not found, try partial name match
            if not employee and employee_name:
                search_domain = base_domain + [('name', 'ilike', employee_name)]
                employee = self.env['hr.employee'].search(search_domain, limit=1)
            
            if not employee:
                errors.append(_("Row %d: Employee '%s' (Code: %s) not found or you don't have access.") % (row_idx, employee_name or 'N/A', employee_code or 'N/A'))
                continue
            
            # Additional check: verify user has access to this employee
            if hasattr(self.env.user, '_get_restricted_employee_ids'):
                allowed_ids = self.env.user._get_restricted_employee_ids()
                if allowed_ids is not None and employee.id not in allowed_ids:
                    errors.append(_("Row %d: You are not authorized to create attendance for employee '%s'.") % (row_idx, employee.name))
                    continue
            
            # Get values
            total_present_days = int(row[2]) if len(row) > 2 and row[2] else 0
            total_overtime = float(row[3]) if len(row) > 3 and row[3] else 0.0
            check_in = float(row[4]) if len(row) > 4 and row[4] else (self.default_check_in or 9.0)
            check_out = float(row[5]) if len(row) > 5 and row[5] else (self.default_check_out or 18.0)
            
            if total_present_days <= 0:
                errors.append(_("Row %d: Employee '%s' - Total Present Days must be greater than 0.") % (row_idx, employee.name))
                continue
            
            lines_to_create.append({
                'wizard_id': self.id,
                'employee_id': employee.id,
                'total_present_days': total_present_days,
                'total_overtime_hours': total_overtime,
                'check_in_hours': check_in,
                'check_out_hours': check_out,
            })
        
        # Create lines
        created_count = 0
        if lines_to_create:
            created_lines = self.env['monthly.attendance.employee.line'].create(lines_to_create)
            created_count = len(created_lines)
        
        message = _('Imported %d employee(s) successfully.') % created_count
        if errors:
            message += '\n\n' + _('Errors:') + '\n' + '\n'.join(errors[:10])  # Show first 10 errors
            if len(errors) > 10:
                message += '\n' + _('... and %d more errors.') % (len(errors) - 10)
        
        # Return form view action to reload and show imported data
        # This will reload the wizard dialog with the new employee lines visible
        return {
            'type': 'ir.actions.act_window',
            'name': _('Create Monthly Attendance'),
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'views': [(False, 'form')],
            'context': dict(self.env.context),
        }


class MonthlyAttendanceEmployeeLine(models.TransientModel):
    """Employee line in the wizard"""
    _name = 'monthly.attendance.employee.line'
    _description = 'Monthly Attendance Employee Line'
    _order = 'employee_id'

    wizard_id = fields.Many2one(
        'monthly.attendance.config.wizard',
        string='Wizard',
        required=True,
        ondelete='cascade',
    )
    @api.model
    def _get_employee_line_domain(self):
        """Get domain for employee field in employee line based on manager restrictions."""
        domain = [('active', '=', True)]
        if hasattr(self.env.user, '_get_restricted_employee_ids'):
            allowed_ids = self.env.user._get_restricted_employee_ids()
            if allowed_ids is not None:
                if not allowed_ids:
                    domain.append(('id', '=', False))  # No access
                else:
                    domain.append(('id', 'in', allowed_ids))
        return domain

    employee_id = fields.Many2one(
        'hr.employee',
        string='Employee',
        required=True,
        domain=lambda self: self._get_employee_line_domain(),
    )
    total_present_days = fields.Integer(
        string='Total Present Days',
        required=True,
        default=0,
    )
    total_overtime_hours = fields.Float(
        string='Total Overtime (Hours)',
        default=0.0,
    )
    check_in_hours = fields.Float(
        string='Check In (Hours)',
    )
    check_out_hours = fields.Float(
        string='Check Out (Hours)',
    )

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        wizard_id = self.env.context.get('default_wizard_id')
        if wizard_id:
            wizard = self.env['monthly.attendance.config.wizard'].browse(wizard_id)
            if 'check_in_hours' in fields_list and not res.get('check_in_hours'):
                res['check_in_hours'] = wizard.default_check_in or 9.0
            if 'check_out_hours' in fields_list and not res.get('check_out_hours'):
                res['check_out_hours'] = wizard.default_check_out or 18.0
        else:
            if 'check_in_hours' in fields_list and not res.get('check_in_hours'):
                res['check_in_hours'] = 9.0
            if 'check_out_hours' in fields_list and not res.get('check_out_hours'):
                res['check_out_hours'] = 18.0
        return res

    def _get_working_weekdays(self, calendar):
        """Return set of Python weekday integers (0=Monday, 6=Sunday) that are working days"""
        if not calendar or not calendar.attendance_ids:
            return {0, 1, 2, 3, 4}  # Monday-Friday
        return {int(a.dayofweek) for a in calendar.attendance_ids}

    def _get_holiday_dates(self, calendar, date_from, date_to):
        """Return set of dates that are holidays"""
        if not calendar:
            return set()
        leaves = self.env['resource.calendar.leaves'].search([
            ('calendar_id', '=', calendar.id),
            ('date_from', '<=', datetime.combine(date_to, datetime.max.time())),
            ('date_to', '>=', datetime.combine(date_from, datetime.min.time())),
        ])
        holiday_dates = set()
        for leave in leaves:
            start_dt = fields.Datetime.to_datetime(leave.date_from)
            end_dt = fields.Datetime.to_datetime(leave.date_to)
            if start_dt and end_dt:
                cur = start_dt.date()
                end = end_dt.date()
                while cur <= end:
                    if date_from <= cur <= date_to:
                        holiday_dates.add(cur)
                    cur += timedelta(days=1)
        return holiday_dates

    def _get_working_days_in_month(self):
        """Return sorted list of working days in the month"""
        wizard = self.wizard_id
        try:
            y = int(wizard.year)
            m = int(wizard.month)
        except (ValueError, TypeError):
            raise UserError(_('Invalid month or year.'))

        last_day = monthrange(y, m)[1]
        date_from = date(y, m, 1)
        date_to = date(y, m, last_day)

        calendar = wizard.resource_calendar_id
        working_weekdays = self._get_working_weekdays(calendar)
        holidays = self._get_holiday_dates(calendar, date_from, date_to)

        working_days = []
        d = date_from
        while d <= date_to:
            if d.weekday() in working_weekdays and d not in holidays:
                working_days.append(d)
            d += timedelta(days=1)
        return sorted(working_days)

    def _float_to_time(self, float_val):
        """Convert float hour to time object"""
        h = int(float_val)
        m = int((float_val - h) * 60)
        return time(h, min(m, 59), 0)

    def _localize_dt(self, dt):
        """Convert naive datetime to UTC"""
        user_tz = self.env.user.tz or 'UTC'
        local = pytz.timezone(user_tz)
        local_dt = local.localize(dt, is_dst=None)
        return local_dt.astimezone(pytz.UTC).replace(tzinfo=None)

    def create_attendance_records(self):
        """Create attendance records for this employee"""
        self.ensure_one()
        # Check if user is allowed to create attendance for this employee
        if hasattr(self.env.user, '_get_restricted_employee_ids'):
            allowed_ids = self.env.user._get_restricted_employee_ids()
            if allowed_ids is not None and self.employee_id.id not in allowed_ids:
                raise UserError(_('You are not authorized to create attendance for employee %s.') % self.employee_id.name)
        
        if self.total_present_days <= 0:
            raise UserError(_('Total Present Days must be greater than 0.'))

        working_days = self._get_working_days_in_month()
        if not working_days:
            raise UserError(_('No working days found in the selected month.'))

        n_days = min(self.total_present_days, len(working_days))
        days_to_mark = working_days[:n_days]

        # Always use 8 hours as base working hours per day (standard)
        base_hours = 8.0
        
        # Calculate how to distribute overtime as whole hours
        total_overtime = self.total_overtime_hours or 0.0
        overtime_days_count = int(total_overtime / 8)  # Full 8-hour overtime days
        remaining_overtime = total_overtime % 8  # Remaining overtime hours

        t_in = self._float_to_time(self.check_in_hours)
        created = 0
        employee = self.employee_id
        user_tz = pytz.timezone(self.env.user.tz or 'UTC')

        for idx, d in enumerate(days_to_mark):
            # Skip if attendance already exists
            local_start = user_tz.localize(datetime.combine(d, time.min))
            local_end = user_tz.localize(datetime.combine(d, time.max))
            utc_start = local_start.astimezone(pytz.UTC).replace(tzinfo=None)
            utc_end = local_end.astimezone(pytz.UTC).replace(tzinfo=None)
            exists = self.env['hr.attendance'].search_count([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', utc_start),
                ('check_in', '<=', utc_end),
            ])
            if exists:
                continue

            # Determine hours for this day
            # Add overtime to the first few days as whole hours
            day_overtime = 0.0
            if idx < overtime_days_count:
                # Full 8-hour overtime day
                day_overtime = 8.0
            elif idx == overtime_days_count and remaining_overtime > 0:
                # Add remaining overtime hours to this day
                day_overtime = remaining_overtime
            
            hours_for_day = base_hours + day_overtime

            dt_in = datetime.combine(d, t_in)
            dt_out = dt_in + timedelta(hours=hours_for_day)
            utc_in = self._localize_dt(dt_in)
            utc_out = self._localize_dt(dt_out)

            self.env['hr.attendance'].create({
                'employee_id': employee.id,
                'check_in': utc_in,
                'check_out': utc_out,
            })
            created += 1

        return created
