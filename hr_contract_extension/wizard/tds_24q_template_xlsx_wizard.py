# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError

import base64
import io
from copy import copy
from datetime import date
from dateutil.relativedelta import relativedelta
import math

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


class Tds24QTemplateXlsxWizard(models.TransientModel):
    _name = 'tds.24q.template.xlsx.wizard'
    _description = 'TDS 24Q Export (Template Based XLSX)'

    date_from = fields.Date(string='From', required=True)
    date_to = fields.Date(string='To', required=True)

    employee_filter = fields.Selection(
        [('all', 'All Employees'), ('selected', 'Selected Employees')],
        string='Employees',
        default='all',
        required=True,
    )
    employee_ids = fields.Many2many('hr.employee', string='Selected Employees')

    template_path = fields.Char(string='Template Path', required=True)

    output_file = fields.Binary('24Q XLSX', readonly=True)
    output_file_name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('draft', 'Configure'), ('done', 'Done')], default='draft')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        icp = self.env['ir.config_parameter'].sudo()
        default_path = r'D:\odoo\odoo-17.0\harman_star_17\kpt_addons\24QsheetKPT425_Ex 20-05-2025.xlsx'
        res.setdefault('template_path', icp.get_param('hr_contract_extension.tds_24q_template_path', default_path))

        today = fields.Date.context_today(self)
        fy_year = today.year if today.month >= 4 else (today.year - 1)
        fy_start = date(fy_year, 4, 1)
        fy_end = date(fy_year + 1, 3, 31)
        res.setdefault('date_from', fy_start)
        res.setdefault('date_to', fy_end)
        return res

    def _style_from_template(self, ws, row, col, template_row):
        tmpl = ws.cell(row=template_row, column=col)
        cell = ws.cell(row=row, column=col)
        try:
            cell._style = copy(tmpl._style)
            cell.number_format = tmpl.number_format
            cell.font = copy(tmpl.font)
            cell.fill = copy(tmpl.fill)
            cell.border = copy(tmpl.border)
            cell.alignment = copy(tmpl.alignment)
            cell.protection = copy(tmpl.protection)
        except Exception:
            pass
        return cell

    def _format_ddmmyyyy(self, dt):
        if not dt:
            return ''
        if isinstance(dt, date):
            return dt.strftime('%d/%m/%Y')
        return str(dt)

    def _challan_detail_text(self, sr_no, challan):
        if not challan:
            return ''
        bsr = challan.bsr_code or ''
        dt = challan.tax_deposit_date
        challan_no = challan.challan_no or ''
        dt_txt = dt.strftime('%d-%b-%Y') if isinstance(dt, date) else ''
        return f"{sr_no}  ({bsr}, {dt_txt}, {challan_no})".strip()

    def _iter_month_starts(self, d1, d2):
        cur = d1.replace(day=1)
        end = d2.replace(day=1)
        while cur <= end:
            yield cur
            cur += relativedelta(months=1)

    def _month_year_to_first_date(self, month_year):
        # month_year example: "April 2024"
        try:
            dt = fields.Date.from_string('01 ' + (month_year or ''))
        except Exception:
            return False
        return dt

    def _monthly_payment_amount_for_employee(self, employee, month_dt):
        # Monthly taxable/payment amount for 24Q: prefer salary increment lines if available (monthly prorated by
        # calendar days + one-time amounts for the month). Fallback to contract proration.
        Tds = self.env['hr.tds']
        tds = Tds.search([
            ('hr_employee_id', '=', employee.id),
            ('tds_from_date', '<=', month_dt),
            ('tds_to_date', '>=', month_dt),
        ], order='tds_from_date desc, id desc', limit=1)

        month_start = month_dt.replace(day=1)
        month_end = month_start + relativedelta(months=1, days=-1)

        def _overlap_days(start_a, end_a, start_b, end_b):
            if not (start_a and end_a and start_b and end_b):
                return 0
            s = max(start_a, start_b)
            e = min(end_a, end_b)
            if s > e:
                return 0
            return (e - s).days + 1

        if tds and tds.salary_increment_ids:
            fy_start = tds.tds_from_date or month_start
            fy_end = tds.tds_to_date or month_end

            total_days = (month_end - month_start).days + 1
            gross_total = 0.0

            monthly_lines = tds.salary_increment_ids.filtered(
                lambda l: l.line_type == 'monthly'
                and (l.effective_from or fy_start) <= month_end
                and ((l.effective_to or fy_end) >= month_start)
            )
            for l in monthly_lines:
                l_from = max(l.effective_from or month_start, month_start, fy_start)
                l_to = min(l.effective_to or month_end, month_end, fy_end)
                od = _overlap_days(l_from, l_to, month_start, month_end)
                if od and total_days:
                    gross_total += float(l.monthly_gross or 0.0) * (float(od) / float(total_days))

            one_time_lines = tds.salary_increment_ids.filtered(
                lambda l: l.line_type == 'one_time'
                and l.effective_from
                and l.effective_from >= fy_start
                and l.effective_from <= fy_end
                and l.effective_from.year == month_dt.year
                and l.effective_from.month == month_dt.month
            )
            gross_total += float(sum(one_time_lines.mapped('one_time_amount')) or 0.0)
            return round(max(float(gross_total or 0.0), 0.0), 2)

        contracts = self.env['hr.contract'].search([
            ('employee_id', '=', employee.id),
            ('state', 'in', ['open', 'close']),
            ('date_start', '<=', month_end),
            '|',
            ('date_end', '=', False),
            ('date_end', '>=', month_start),
        ], order='date_start')
        if not contracts:
            return 0.0

        chosen = contracts[-1]
        monthly_salary = chosen.gross if getattr(chosen, 'gross', 0.0) else (chosen.wage or 0.0)
        monthly_salary = float(monthly_salary or 0.0)

        eff_start = max(month_start, chosen.date_start)
        eff_end = min(month_end, chosen.date_end) if chosen.date_end else month_end

        total_days = (month_end - month_start).days + 1
        eff_days = _overlap_days(eff_start, eff_end, month_start, month_end)
        prorate = (float(eff_days) / float(total_days)) if total_days else 0.0
        return round(max(monthly_salary * prorate, 0.0), 2)

    def _get_employees_for_report(self):
        self.ensure_one()
        if self.employee_filter == 'selected':
            return self.employee_ids

        # All Employees: include employees who have a TDS record overlapping the requested range.
        tds_recs = self.env['hr.tds'].search([
            ('tds_from_date', '<=', self.date_to),
            ('tds_to_date', '>=', self.date_from),
        ])
        return tds_recs.mapped('hr_employee_id').filtered(lambda e: e)

    def _tds_for_employee_in_range(self, employee):
        self.ensure_one()
        Tds = self.env['hr.tds']
        # Prefer the TDS record used on payslip (is_tds_payslip) if present; fallback to latest in range.
        return Tds.search([
            ('hr_employee_id', '=', employee.id),
            ('tds_from_date', '<=', self.date_to),
            ('tds_to_date', '>=', self.date_from),
        ], order='is_tds_payslip desc, tds_from_date desc, id desc', limit=1)

    def action_generate(self):
        self.ensure_one()

        if not load_workbook:
            raise UserError(
                "Required python dependency for XLSX generation is missing.\n"
                "Install: pip install openpyxl\n"
                "Then restart Odoo server."
            )

        if not self.template_path:
            raise UserError('Template path is required.')

        icp = self.env['ir.config_parameter'].sudo()
        if self.template_path != icp.get_param('hr_contract_extension.tds_24q_template_path'):
            icp.set_param('hr_contract_extension.tds_24q_template_path', self.template_path)

        try:
            wb = load_workbook(self.template_path)
        except Exception as e:
            raise UserError(f"Cannot open template XLSX at: {self.template_path}\nError: {e}")

        required_sheets = {'Challan', 'Challan_adjustment', 'Employee_Salary_Detail'}
        if not required_sheets.issubset(set(wb.sheetnames)):
            raise UserError("Template must contain sheets: Challan, Challan_adjustment, Employee_Salary_Detail")

        ws_challan = wb['Challan']
        ws_adj = wb['Challan_adjustment']
        ws_emp = wb['Employee_Salary_Detail']

        company = self.env.company
        tan_no = company.vat or ''

        # Challan sheet header
        ws_challan['B1'].value = tan_no
        ws_challan.cell(row=1, column=20).value = self.date_from
        ws_challan.cell(row=1, column=21).value = self.date_to

        # --- Clear existing data rows (keep formulas/extra columns intact) ---
        def _clear_cells(ws, start_row, max_col):
            for r in range(start_row, ws.max_row + 1):
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).value = None

        _clear_cells(ws_challan, start_row=5, max_col=14)
        _clear_cells(ws_adj, start_row=4, max_col=20)
        _clear_cells(ws_emp, start_row=5, max_col=102)

        # ==================== SHEET 1: Challan ====================
        m_from = self.date_from.replace(day=1)
        m_to = self.date_to.replace(day=1)

        employees = self._get_employees_for_report()
        # Stabilize ordering
        employees = employees.sorted(lambda e: (e.name or '', e.id))

        emp_tds_map = {}
        for emp in employees:
            tds = self._tds_for_employee_in_range(emp)
            if tds:
                # Ensure month-wise lines are up to date before exporting.
                # This prevents stale rows / duplicates when month_ids were not refreshed.
                try:
                    tds._sync_month_ids_lines()
                except Exception:
                    pass
                emp_tds_map[emp.id] = tds

        # Prefer period-based filtering (one challan per month). Fallback to tax_deposit_date
        # for older challans where period_month might be empty.
        challans = self.env['tds.challan'].search([
            '|',
            '&',
            ('period_month', '!=', False),
            ('period_month', '>=', m_from),
            ('period_month', '<=', m_to),
            '&',
            ('period_month', '=', False),
            ('tax_deposit_date', '>=', self.date_from),
            ('tax_deposit_date', '<=', self.date_to),
        ], order='period_month, tax_deposit_date, id')

        challan_row_start = 5
        challan_template_row = challan_row_start
        challan_sno_map = {}

        for idx, ch in enumerate(challans, start=1):
            r = challan_row_start + (idx - 1)
            challan_sno_map[ch.id] = idx

            self._style_from_template(ws_challan, r, 1, challan_template_row).value = idx
            self._style_from_template(ws_challan, r, 2, challan_template_row).value = float(ch.tds_payment or 0.0)
            self._style_from_template(ws_challan, r, 3, challan_template_row).value = float(ch.surcharge or 0.0)
            self._style_from_template(ws_challan, r, 4, challan_template_row).value = float(ch.education_cess or 0.0)
            self._style_from_template(ws_challan, r, 5, challan_template_row).value = float(ch.higher_education_cess or 0.0)
            self._style_from_template(ws_challan, r, 6, challan_template_row).value = float(ch.interest or 0.0)
            self._style_from_template(ws_challan, r, 7, challan_template_row).value = float(ch.other or 0.0)
            self._style_from_template(ws_challan, r, 8, challan_template_row).value = float(ch.fee or 0.0)
            self._style_from_template(ws_challan, r, 9, challan_template_row).value = ch.cheque_dd_no or ''
            self._style_from_template(ws_challan, r, 10, challan_template_row).value = ch.bsr_code or ''
            self._style_from_template(ws_challan, r, 11, challan_template_row).value = ch.tax_deposit_date
            self._style_from_template(ws_challan, r, 12, challan_template_row).value = ch.challan_no or ''
            self._style_from_template(ws_challan, r, 13, challan_template_row).value = 'Yes' if ch.book_entry == 'yes' else 'No'
            self._style_from_template(ws_challan, r, 14, challan_template_row).value = ch.minor_head or ''

        # ==================== SHEET 2: Challan_adjustment ====================
        adj_row_start = 4
        adj_template_row = adj_row_start

        def _employee_ref(emp):
            return getattr(emp, 'employee_code', False) or getattr(emp, 'barcode', False) or getattr(emp, 'identification_id', False) or ''

        # Build Challan_adjustment rows from the selected employee -> TDS mapping,
        # ensuring one authoritative TDS record per employee (prevents duplicate rows).
        rows = []
        for emp in employees:
            tds = emp_tds_map.get(emp.id)
            if not tds:
                continue
            eff_from = max(self.date_from, tds.tds_from_date) if (self.date_from and tds.tds_from_date) else (self.date_from or tds.tds_from_date)
            eff_to = min(self.date_to, tds.tds_to_date) if (self.date_to and tds.tds_to_date) else (self.date_to or tds.tds_to_date)
            if not eff_from or not eff_to or eff_from > eff_to:
                continue
            for month_dt in self._iter_month_starts(eff_from, eff_to):
                month_key = f"{month_dt.strftime('%B')} {month_dt.year}".strip().lower()
                line = tds.month_ids.filtered(
                    lambda l: (l.tds_month_year or '').strip().lower() == month_key and not l.is_previous_employer
                )
                line = line[:1]
                if not line:
                    continue
                tds_amt = float(line[0].tds_month_amt or 0.0)
                if not tds_amt:
                    continue
                dt = month_dt.replace(day=1)
                rows.append((dt, emp, tds, line[0]))

        rows.sort(key=lambda x: (x[0], x[1].id, x[3].id))

        for i, (dt, emp, tds, line) in enumerate(rows, start=1):
            r = adj_row_start + (i - 1)

            month_end = dt.replace(day=1) + relativedelta(months=1, days=-1)
            taxable_amt = self._monthly_payment_amount_for_employee(emp, month_end)

            month_start = dt.replace(day=1)
            challan = self.env['tds.challan'].search([
                ('company_id', '=', emp.company_id.id),
                ('period_month', '=', month_start),
            ], order='tax_deposit_date, id', limit=1)
            challan_sno = challan_sno_map.get(challan.id) if challan else False

            # TDS amount for 24Q should include any extra one-time TDS for bonus/arrears
            # (added on top of standard monthly TDS).
            tds_amt = float(line.tds_month_amt or 0.0)

            extra_tds = 0.0
            if tds and getattr(tds, 'salary_increment_ids', False):
                extra_lines = tds.salary_increment_ids.filtered(
                    lambda l: l.line_type == 'one_time'
                    and l.effective_from
                    and l.effective_from.year == dt.year
                    and l.effective_from.month == dt.month
                    and (l.one_time_tds or 0.0)
                    and (l.one_time_tds or 0.0) != 0.0
                )
                extra_tds = float(sum(extra_lines.mapped('one_time_tds')) or 0.0)

            # Avoid double counting: only add extra when month_ids amount looks like base monthly
            # (i.e., it hasn't already absorbed the extra one-time TDS).
            base_monthly = float(getattr(tds, 'tds_deduction_month_cess', 0.0) or 0.0) if tds else 0.0
            if extra_tds and base_monthly and abs(float(tds_amt) - float(base_monthly)) < 0.01:
                tds_amt = float(tds_amt) + float(extra_tds)
            surcharge = 0.0
            edu_cess = 0.0
            higher_cess = 0.0
            total_tax = tds_amt + surcharge + edu_cess + higher_cess

            challan_detail = self._challan_detail_text(challan_sno or 0, challan)

            # A..T (1..20)
            self._style_from_template(ws_adj, r, 1, adj_template_row).value = i
            self._style_from_template(ws_adj, r, 2, adj_template_row).value = _employee_ref(emp)
            self._style_from_template(ws_adj, r, 3, adj_template_row).value = getattr(emp, 'pan_number', False) or ''
            self._style_from_template(ws_adj, r, 4, adj_template_row).value = emp.name or ''
            self._style_from_template(ws_adj, r, 5, adj_template_row).value = month_end
            self._style_from_template(ws_adj, r, 6, adj_template_row).value = None
            self._style_from_template(ws_adj, r, 7, adj_template_row).value = float(taxable_amt or 0.0)
            self._style_from_template(ws_adj, r, 8, adj_template_row).value = tds_amt
            self._style_from_template(ws_adj, r, 9, adj_template_row).value = surcharge
            self._style_from_template(ws_adj, r, 10, adj_template_row).value = edu_cess
            self._style_from_template(ws_adj, r, 11, adj_template_row).value = higher_cess
            self._style_from_template(ws_adj, r, 12, adj_template_row).value = total_tax
            self._style_from_template(ws_adj, r, 13, adj_template_row).value = total_tax
            self._style_from_template(ws_adj, r, 14, adj_template_row).value = month_end
            self._style_from_template(ws_adj, r, 15, adj_template_row).value = challan.tax_deposit_date if challan else None
            self._style_from_template(ws_adj, r, 16, adj_template_row).value = challan.challan_no if challan else ''
            self._style_from_template(ws_adj, r, 17, adj_template_row).value = challan.bsr_code if challan else ''
            self._style_from_template(ws_adj, r, 18, adj_template_row).value = challan_detail
            self._style_from_template(ws_adj, r, 19, adj_template_row).value = ''
            self._style_from_template(ws_adj, r, 20, adj_template_row).value = ''

        # ==================== SHEET 3: Employee_Salary_Detail ====================
        # One row per employee based on the selected authoritative TDS record.
        tds_records = self.env['hr.tds'].browse([t.id for t in emp_tds_map.values()])
        tds_records = tds_records.sorted(lambda t: (t.hr_employee_id.name or '', t.hr_employee_id.id, t.id))

        emp_row_start = 5
        emp_template_row = emp_row_start

        category_label = {
            'W': 'Woman',
            'S': 'Senior Citizen',
            'O': 'Super Senior Citizen',
            'G': 'Other',
        }

        for idx, tds in enumerate(tds_records, start=1):
            emp = tds.hr_employee_id
            if not emp:
                continue

            contract = tds.hr_contract_id
            prof_tax_annual = (getattr(contract, 'professional_tax', 0.0) or 0.0) * 12.0 if contract else 0.0

            section_totals = {}
            for line in tds.deduction_ids:
                sec = (line.section_id.name or '').strip().upper()
                section_totals[sec] = section_totals.get(sec, 0.0) + (line.deduction_amt or 0.0)

            code = getattr(emp, 'tds_category_code', False) or 'G'
            cat_txt = f"{code} - {category_label.get(code, 'Other')}"

            designation = getattr(emp, 'job_title', False) or (emp.job_id.name if getattr(emp, 'job_id', False) else '') or ''

            # Build values list in the same order as the government template columns.
            values = []
            values.append(getattr(emp, 'pan_number', False) or '')
            values.append(emp.name or '')
            values.append(designation)
            values.append(cat_txt)
            values.append(self._format_ddmmyyyy(tds.tds_from_date))
            values.append(self._format_ddmmyyyy(tds.tds_to_date))

            annual_salary = float(tds.annual_salary or 0.0)
            perq = float(getattr(tds, 'perquisites_17_2', 0.0) or 0.0)
            profits = float(getattr(tds, 'profits_17_3', 0.0) or 0.0)

            values.append(annual_salary)
            values.append(perq)
            values.append(profits)
            values.append(annual_salary + perq + profits)
            values.append(float(getattr(tds, 'previous_employer_taxable', 0.0) or 0.0))

            values.append(float(getattr(tds, 'travel_concession_10_5', 0.0) or 0.0))
            values.append(float(getattr(tds, 'gratuity_10_10', 0.0) or 0.0))
            values.append(float(getattr(tds, 'commuted_pension_10_10a', 0.0) or 0.0))
            values.append(float(getattr(tds, 'leave_salary_10_10aa', 0.0) or 0.0))
            values.append(float(getattr(tds, 'final_hra_amt', 0.0) or 0.0))
            values.append(float(getattr(tds, 'other_exemptions_10', 0.0) or 0.0))
            values.append(float(getattr(tds, 'total_exemptions_10', 0.0) or 0.0))
            values.append(float(getattr(tds, 'standard_deduction_16_ii', 0.0) or 0.0))
            values.append(float(prof_tax_annual or 0.0))

            values.append(float(getattr(tds, 'income_chargeable_salaries', 0.0) or 0.0))
            values.append(0.0)
            values.append(0.0)
            values.append(float(getattr(tds, 'taxable_amount', 0.0) or 0.0))

            # VI-A sections (gross/deductible pairs - we fill same where applicable)
            v_80c = float(section_totals.get('80C', 0.0) or 0.0)
            v_80ccc = float(section_totals.get('80CCC', 0.0) or 0.0)
            v_80ccd1 = float(section_totals.get('80CCD(1)', 0.0) or 0.0)
            v_80ccd1b = float(section_totals.get('80CCD(1B)', 0.0) or 0.0)
            v_80ccd2 = float(section_totals.get('80CCD(2)', 0.0) or 0.0)
            v_80d = float(section_totals.get('80D', 0.0) or 0.0)
            v_80e = float(section_totals.get('80E', 0.0) or 0.0)
            v_80g = float(section_totals.get('80G', 0.0) or 0.0)
            v_80tta = float(section_totals.get('80TTA', 0.0) or 0.0)

            values += [v_80c, v_80c]
            values += [v_80ccc, v_80ccc]
            values += [v_80ccd1, v_80ccd1]
            values += [min(150000.0, v_80c + v_80ccc + v_80ccd1)]
            values += [v_80ccd1b, v_80ccd1b]
            values += [v_80ccd2, v_80ccd2]
            values += [v_80d, v_80d]
            values += [v_80e, v_80e]
            values += [v_80g, v_80g, v_80g]
            values += [v_80tta, v_80tta, v_80tta]

            # Other VI-A placeholders (gross/qualifying/deductible) + totals
            values += [0.0, 0.0, 0.0]
            values += [float(getattr(tds, 'total_deductions', 0.0) or 0.0)]

            values.append(float(getattr(tds, 'taxable_amount', 0.0) or 0.0))
            values.append(float(getattr(tds, 'tax_payable', 0.0) or 0.0))
            values.append(float(getattr(tds, 'rebate_87a', 0.0) or 0.0))
            values.append(float(getattr(tds, 'surcharge', 0.0) or 0.0))
            values.append(float(getattr(tds, 'health_education_cess', 0.0) or 0.0))
            values.append(0.0)
            values.append(float(getattr(tds, 'tax_payable_cess', 0.0) or 0.0))

            values.append(float(getattr(tds, 'tax_payable_cess', 0.0) or 0.0))
            values.append(0.0)
            values.append(float(getattr(tds, 'tax_payable_cess', 0.0) or 0.0))
            values.append(0.0)

            values.append('Yes' if (float(getattr(tds, 'final_hra_amt', 0.0) or 0.0) > 100000.0) else 'No')

            # Landlord / lender / superannuation blocks - keep as in previous export approach
            for i in range(13):
                values.append('' if i % 3 == 0 else ('0' if i == 1 else ''))
            values.append('No')
            for i in range(13):
                values.append('' if i % 3 == 0 else ('0' if i == 1 else ''))
            values.append('No')

            values += ['', '', '']
            values += [0.0, 0.0, 0.0, 0.0]
            values.append('Yes' if bool(getattr(tds, 'new_tax_regime_opted', False)) else 'No')
            values += [0.0, 0.0, '', '', '']

            # Ensure exactly 102 columns
            if len(values) < 102:
                values += [''] * (102 - len(values))
            elif len(values) > 102:
                values = values[:102]

            r = emp_row_start + (idx - 1)
            for c, v in enumerate(values, start=1):
                self._style_from_template(ws_emp, r, c, emp_template_row).value = v

        output = io.BytesIO()
        try:
            wb.save(output)
        except Exception as e:
            raise UserError(f"Failed to generate XLSX: {e}")

        output.seek(0)
        data = output.getvalue()
        output.close()

        filename = f"24Q_Report_{fields.Date.to_string(self.date_from)}_to_{fields.Date.to_string(self.date_to)}.xlsx"
        self.write({
            'output_file': base64.b64encode(data),
            'output_file_name': filename,
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_download(self):
        self.ensure_one()
        if not self.output_file:
            raise UserError('No file available. Please Generate first.')
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model={self._name}&id={self.id}&field=output_file&download=true&filename={self.output_file_name}',
            'target': 'self',
        }

    def action_new(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }
