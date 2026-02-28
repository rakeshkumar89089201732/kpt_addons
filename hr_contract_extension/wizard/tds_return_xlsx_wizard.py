# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError

import base64
import io
from copy import copy
import re
from datetime import date
from dateutil.relativedelta import relativedelta

try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None
    Workbook = None


class TdsReturnXlsxWizard(models.TransientModel):
    _name = 'tds.return.xlsx.wizard'
    _description = 'TDS Return XLSX (Template Based)'

    fy_start_year = fields.Integer(string='Financial Year (Start Year)', required=True)
    quarter = fields.Selection(
        [('q1', 'Quarter 1 (Apr-Jun)'), ('q2', 'Quarter 2 (Jul-Sep)'), ('q3', 'Quarter 3 (Oct-Dec)'), ('q4', 'Quarter 4 (Jan-Mar)')],
        string='Quarter',
        required=True,
    )

    date_from = fields.Date(string='From', required=True)
    date_to = fields.Date(string='To', required=True)

    employee_filter = fields.Selection(
        [('all', 'All Employees'), ('selected', 'Selected Employees')],
        string='Employees',
        default='all',
        required=True,
    )
    employee_ids = fields.Many2many('hr.employee', string='Selected Employees')

    challan_id = fields.Many2one('tds.challan', string='Challan (for Annexure)', help='Optional: If set, Annexure rows will use this challan details.')

    section_code = fields.Char(string='Section Code', default='192', required=True)

    template_path = fields.Char(string='Template Path', required=True)

    output_file = fields.Binary('TDS Return XLSX', readonly=True)
    output_file_name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('draft', 'Configure'), ('done', 'Done')], default='draft')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        icp = self.env['ir.config_parameter'].sudo()
        default_path = r'D:\odoo\odoo-17.0\harman_star_17\xxf26QdtridKPT (1).xlsx'
        res.setdefault('template_path', icp.get_param('hr_contract_extension.tds_return_template_path', default_path))

        today = fields.Date.context_today(self)
        fy_year = today.year if today.month >= 4 else (today.year - 1)
        if today.month in (4, 5, 6):
            q = 'q1'
        elif today.month in (7, 8, 9):
            q = 'q2'
        elif today.month in (10, 11, 12):
            q = 'q3'
        else:
            q = 'q4'

        res.setdefault('fy_start_year', fy_year)
        res.setdefault('quarter', q)
        q_from, q_to = self._quarter_dates(fy_year, q)
        res.setdefault('date_from', q_from)
        res.setdefault('date_to', q_to)
        return res

    def _quarter_dates(self, fy_start_year, quarter):
        fy_start = date(int(fy_start_year), 4, 1)
        fy_end = date(int(fy_start_year) + 1, 3, 31)
        if quarter == 'q1':
            d_from = fy_start
            d_to = date(int(fy_start_year), 6, 30)
        elif quarter == 'q2':
            d_from = date(int(fy_start_year), 7, 1)
            d_to = date(int(fy_start_year), 9, 30)
        elif quarter == 'q3':
            d_from = date(int(fy_start_year), 10, 1)
            d_to = date(int(fy_start_year), 12, 31)
        else:
            d_from = date(int(fy_start_year) + 1, 1, 1)
            d_to = fy_end
        return d_from, d_to

    @api.onchange('fy_start_year', 'quarter')
    def _onchange_quarter_dates(self):
        for wiz in self:
            if not wiz.fy_start_year or not wiz.quarter:
                continue
            d_from, d_to = wiz._quarter_dates(wiz.fy_start_year, wiz.quarter)
            wiz.date_from = d_from
            wiz.date_to = d_to

    def _get_employees_for_report(self):
        self.ensure_one()
        if self.employee_filter == 'selected':
            return self.employee_ids

        tds_recs = self.env['hr.tds'].search([
            ('tds_from_date', '<=', self.date_to),
            ('tds_to_date', '>=', self.date_from),
        ])
        return tds_recs.mapped('hr_employee_id').filtered(lambda e: e)

    def _tds_for_employee_in_range(self, employee):
        self.ensure_one()
        return self.env['hr.tds'].search([
            ('hr_employee_id', '=', employee.id),
            ('tds_from_date', '<=', self.date_to),
            ('tds_to_date', '>=', self.date_from),
        ], order='is_tds_payslip desc, tds_from_date desc, id desc', limit=1)

    def action_generate(self):
        self.ensure_one()

        # Enforce quarterly dates from FY + quarter selection.
        if self.fy_start_year and self.quarter:
            q_from, q_to = self._quarter_dates(self.fy_start_year, self.quarter)
            self.date_from = q_from
            self.date_to = q_to

        if not load_workbook:
            raise UserError(
                "Required python dependency for XLSX generation is missing.\n"
                "Install: pip install openpyxl\n"
                "Then restart Odoo server."
            )

        if not self.template_path:
            raise UserError('Template path is required.')

        icp = self.env['ir.config_parameter'].sudo()
        if self.template_path != icp.get_param('hr_contract_extension.tds_return_template_path'):
            icp.set_param('hr_contract_extension.tds_return_template_path', self.template_path)

        import os
        try:
            wb = load_workbook(self.template_path)
        except Exception as e:
            # If file does not exist, allow minimal fallback; otherwise fail fast to enforce correct template usage
            if not os.path.exists(self.template_path):
                if not Workbook:
                    raise UserError(f"Cannot open template XLSX at: {self.template_path}\nError: {e}")
                wb = Workbook()
                # Remove default 'Sheet'
                try:
                    default_ws = wb.active
                    wb.remove(default_ws)
                except Exception:
                    pass
                ws_challan = wb.create_sheet('Challan')
                ws_annexure = wb.create_sheet('Annexure')
                headers_challan = [
                    'Sr No', 'Section', 'TDS', 'Surcharge', 'Edu Cess', 'Higher Edu Cess',
                    'Interest', 'Other', 'Fee', 'Cheque/DD No', 'BSR Code', 'Tax Deposit Date',
                    'Challan No', 'Book Entry', 'Minor Head'
                ]
                for col, title in enumerate(headers_challan, start=1):
                    ws_challan.cell(row=4, column=col).value = title
                headers_annex = [
                    'Deductee Sr', 'PAN', 'First Name', 'Middle Name', 'Last Name', 'Addr1', 'Addr2',
                    'State', 'PIN', 'Amount Paid', 'Date of Payment', 'Section', 'Rate%', 'TDS',
                    'Surcharge', 'Edu Cess', 'HE Cess', 'Date of Deduction', 'Date of Deposit',
                    'BSR Code', 'Challan No', 'Challan Detail', 'TAN Ref No', 'State Code',
                    'Book Entry', 'Remarks', 'Employee Ref'
                ]
                for col, title in enumerate(headers_annex, start=1):
                    ws_annexure.cell(row=4, column=col).value = title
            else:
                raise UserError(
                    "Template file found but is not a valid .xlsx for openpyxl. "
                    "Please open it in Excel and Save As .xlsx (or .xlsm) and update the Template Path.\n"
                    f"Path: {self.template_path}\nOriginal error: {e}"
                )

        if 'Challan' not in wb.sheetnames or 'Annexure' not in wb.sheetnames:
            raise UserError("Template must contain sheets 'Challan' and 'Annexure'.")

        ws_challan = wb['Challan']
        ws_annexure = wb['Annexure']

        company = self.env.company
        party_name = company.name or ''
        tan_no = company.vat or ''

        # Header values
        ws_challan['C1'].value = party_name
        ws_challan['G1'].value = tan_no
        ws_annexure['C1'].value = party_name
        ws_annexure['G1'].value = tan_no

        # Fill challan rows.
        # Prefer period_month for quarter-wise challan selection; fallback to tax_deposit_date.
        m_from = self.date_from.replace(day=1)
        m_to = self.date_to.replace(day=1)
        challans = self.env['tds.challan'].search([
            '|',
            '&',
            ('period_month', '!=', False),
            ('period_month', '>=', m_from),
            ('period_month', '<=', m_to),
            '&',
            ('period_month', '=', False),
            ('tax_deposit_date', '>=', self.date_from),
            ('tax_deposit_date', '<=', self.date_to),
        ], order='tax_deposit_date, id')

        challan_start_row = 5
        template_challan_row = challan_start_row

        def _style_from_template(ws, row, col, template_row):
            tmpl = ws.cell(row=template_row, column=col)
            cell = ws.cell(row=row, column=col)
            # Preserve style from template cell (if any) for newly created rows.
            try:
                cell._style = copy(tmpl._style)
                cell.number_format = tmpl.number_format
                cell.font = copy(tmpl.font)
                cell.fill = copy(tmpl.fill)
                cell.border = copy(tmpl.border)
                cell.alignment = copy(tmpl.alignment)
                cell.protection = copy(tmpl.protection)
            except Exception:
                pass
            return cell

        for idx, ch in enumerate(challans, start=1):
            r = challan_start_row + (idx - 1)
            # Columns: A..O
            _style_from_template(ws_challan, r, 1, template_challan_row).value = idx
            _style_from_template(ws_challan, r, 2, template_challan_row).value = (self.section_code or '').strip()
            _style_from_template(ws_challan, r, 3, template_challan_row).value = float(ch.tds_payment or 0.0)
            _style_from_template(ws_challan, r, 4, template_challan_row).value = float(ch.surcharge or 0.0)
            _style_from_template(ws_challan, r, 5, template_challan_row).value = float(ch.education_cess or 0.0)
            _style_from_template(ws_challan, r, 6, template_challan_row).value = float(ch.higher_education_cess or 0.0)
            _style_from_template(ws_challan, r, 7, template_challan_row).value = float(ch.interest or 0.0)
            _style_from_template(ws_challan, r, 8, template_challan_row).value = float(ch.other or 0.0)
            _style_from_template(ws_challan, r, 9, template_challan_row).value = float(ch.fee or 0.0)
            _style_from_template(ws_challan, r, 10, template_challan_row).value = ch.cheque_dd_no or ''
            _style_from_template(ws_challan, r, 11, template_challan_row).value = ch.bsr_code or ''
            _style_from_template(ws_challan, r, 12, template_challan_row).value = ch.tax_deposit_date
            _style_from_template(ws_challan, r, 13, template_challan_row).value = ch.challan_no or ''
            _style_from_template(ws_challan, r, 14, template_challan_row).value = 'Yes' if ch.book_entry == 'yes' else 'No'
            _style_from_template(ws_challan, r, 15, template_challan_row).value = ch.minor_head or ''

        # Pick challan for annexure details
        challan_for_annexure = self.challan_id
        if not challan_for_annexure and challans:
            challan_for_annexure = challans[0]

        def _iter_month_starts(d1, d2):
            cur = d1.replace(day=1)
            end = d2.replace(day=1)
            while cur <= end:
                yield cur
                cur += relativedelta(months=1)

        def _norm_state_name(name):
            s = (name or '').strip().lower()
            s = s.replace('&', 'and')
            s = re.sub(r'[^a-z0-9\s]', ' ', s)
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        def _build_state_map_from_template(ws):
            # Template has state list in column AF (32) like: "01 - Andaman & Nicobar".
            state_map = {}
            col = 32
            for r in range(1, 200):
                v = ws.cell(row=r, column=col).value
                if not v:
                    continue
                txt = str(v).strip()
                if '-' not in txt:
                    continue
                code, name = txt.split('-', 1)
                code = code.strip()
                name = name.strip()
                if code and name:
                    state_map[_norm_state_name(name)] = f"{code} - {name}"
            return state_map

        template_state_map = _build_state_map_from_template(ws_annexure)

        def _month_line_amount(tds_rec, month_dt):
            if not tds_rec.month_ids:
                return 0.0
            mname = month_dt.strftime('%B').lower()
            year = str(month_dt.year)
            lines = tds_rec.month_ids.filtered(lambda l: (l.months or '') == mname and (l.tds_month_year or '').endswith(year))
            if not lines:
                return 0.0
            # Prefer current employer months
            lines = lines.filtered(lambda l: not l.is_previous_employer) or lines
            return float(lines[0].tds_month_amt or 0.0)

        def _monthly_payment_amount(tds_rec, month_dt):
            # Best-effort monthly gross: use salary increments if present, otherwise contract gross/wage.
            salary_increments = tds_rec.salary_increment_ids
            fy_start = tds_rec.tds_from_date
            fy_end = tds_rec.tds_to_date

            if salary_increments:
                month_start = month_dt.replace(day=1)
                month_end = month_start + relativedelta(months=1, days=-1)

                monthly_lines = salary_increments.filtered(
                    lambda l: l.line_type == 'monthly'
                    and (l.effective_from or fy_start) <= month_end
                    and ((l.effective_to or fy_end) >= month_start)
                )
                monthly_gross = 0.0
                if monthly_lines:
                    chosen = monthly_lines.sorted(lambda l: l.effective_from or fy_start)[-1]
                    monthly_gross = float(chosen.monthly_gross or 0.0)

                one_time_lines = salary_increments.filtered(
                    lambda l: l.line_type == 'one_time'
                    and l.effective_from
                    and l.effective_from.year == month_dt.year
                    and l.effective_from.month == month_dt.month
                )
                one_time_amt = float(sum(one_time_lines.mapped('one_time_amount')) or 0.0)
                return float(monthly_gross) + float(one_time_amt)

            employee = tds_rec.hr_employee_id or (tds_rec.hr_contract_id.employee_id if tds_rec.hr_contract_id else False)
            if employee and month_dt:
                month_start = month_dt.replace(day=1)
                month_end = month_start + relativedelta(months=1, days=-1)
                contracts = self.env['hr.contract'].search([
                    ('employee_id', '=', employee.id),
                    ('state', 'in', ['open', 'close']),
                    ('date_start', '<=', month_end),
                    '|',
                    ('date_end', '=', False),
                    ('date_end', '>=', month_start),
                ], order='date_start')
                if contracts:
                    chosen = contracts[-1]
                    monthly_salary = chosen.gross if getattr(chosen, 'gross', 0.0) else (chosen.wage or 0.0)
                    return float(monthly_salary or 0.0)

            # Fallback
            return float(tds_rec.current_ctc or 0.0) / 12.0 if tds_rec.current_ctc else 0.0

        def _split_name(full_name):
            parts = (full_name or '').strip().split()
            if not parts:
                return '', '', ''
            if len(parts) == 1:
                return parts[0], '', ''
            if len(parts) == 2:
                return parts[0], '', parts[1]
            return parts[0], ' '.join(parts[1:-1]), parts[-1]

        def _employee_address(emp):
            addr = emp.address_home_id if getattr(emp, 'address_home_id', False) else False
            if not addr:
                return '', '', '', ''
            a1 = (addr.street or '')
            a2 = ' '.join([p for p in [addr.street2, addr.city] if p])
            state = addr.state_id.name if addr.state_id else ''
            pin = addr.zip or ''
            return a1, a2, state, pin

        def _challan_detail_text(sr_no, challan):
            if not challan:
                return ''
            bsr = challan.bsr_code or ''
            dt = challan.tax_deposit_date
            challan_no = challan.challan_no or ''
            if isinstance(dt, date):
                dt_txt = dt.strftime('%d-%b-%Y')
            else:
                dt_txt = ''
            return f"{sr_no}  ({bsr}, {dt_txt}, {challan_no})".strip()

        employees = self._get_employees_for_report().sorted(lambda e: (e.name or '', e.id))
        emp_tds_map = {}
        for emp in employees:
            tds = self._tds_for_employee_in_range(emp)
            if tds:
                try:
                    tds._sync_month_ids_lines()
                except Exception:
                    pass
                emp_tds_map[emp.id] = tds

        tds_records = self.env['hr.tds'].browse([t.id for t in emp_tds_map.values()])
        tds_records = tds_records.sorted(lambda t: (t.hr_employee_id.name or '', t.hr_employee_id.id, t.id))

        annexure_start_row = 5
        template_annexure_row = annexure_start_row

        row_idx = 0
        deductee_sr = 1
        for tds in tds_records:
            emp = tds.hr_employee_id
            if not emp:
                continue

            eff_from = max(self.date_from, tds.tds_from_date) if (self.date_from and tds.tds_from_date) else (self.date_from or tds.tds_from_date)
            eff_to = min(self.date_to, tds.tds_to_date) if (self.date_to and tds.tds_to_date) else (self.date_to or tds.tds_to_date)
            if not eff_from or not eff_to or eff_from > eff_to:
                continue

            # Iterate months in quarter (intersection with this employee TDS period)
            for month_dt in _iter_month_starts(eff_from, eff_to):
                tds_amt = _month_line_amount(tds, month_dt)
                if not tds_amt:
                    continue

                pay_amt = _monthly_payment_amount(tds, month_dt)
                rate = (float(tds_amt) / float(pay_amt) * 100.0) if pay_amt else 0.0

                first, middle, last = _split_name(emp.name)
                addr1, addr2, state_name, pin = _employee_address(emp)

                # Prefer template's expected "NN - State" format.
                state_code_name = template_state_map.get(_norm_state_name(state_name), state_name)

                challan_text = _challan_detail_text(1, challan_for_annexure)

                r = annexure_start_row + row_idx
                # A..AA (1..27)
                _style_from_template(ws_annexure, r, 1, template_annexure_row).value = str(deductee_sr)
                _style_from_template(ws_annexure, r, 2, template_annexure_row).value = getattr(emp, 'pan_number', False) or getattr(emp, 'pan_no', False) or ''
                _style_from_template(ws_annexure, r, 3, template_annexure_row).value = first
                _style_from_template(ws_annexure, r, 4, template_annexure_row).value = middle
                _style_from_template(ws_annexure, r, 5, template_annexure_row).value = last
                _style_from_template(ws_annexure, r, 6, template_annexure_row).value = addr1
                _style_from_template(ws_annexure, r, 7, template_annexure_row).value = addr2
                _style_from_template(ws_annexure, r, 8, template_annexure_row).value = state_code_name
                _style_from_template(ws_annexure, r, 9, template_annexure_row).value = pin
                _style_from_template(ws_annexure, r, 10, template_annexure_row).value = float(pay_amt or 0.0)
                _style_from_template(ws_annexure, r, 11, template_annexure_row).value = month_dt
                _style_from_template(ws_annexure, r, 12, template_annexure_row).value = (self.section_code or '').strip()
                _style_from_template(ws_annexure, r, 13, template_annexure_row).value = round(rate, 2)
                _style_from_template(ws_annexure, r, 14, template_annexure_row).value = float(tds_amt or 0.0)
                _style_from_template(ws_annexure, r, 15, template_annexure_row).value = 0.0
                _style_from_template(ws_annexure, r, 16, template_annexure_row).value = 0.0
                _style_from_template(ws_annexure, r, 17, template_annexure_row).value = 0.0
                _style_from_template(ws_annexure, r, 18, template_annexure_row).value = month_dt
                _style_from_template(ws_annexure, r, 19, template_annexure_row).value = challan_for_annexure.tax_deposit_date if challan_for_annexure else None
                _style_from_template(ws_annexure, r, 20, template_annexure_row).value = challan_for_annexure.bsr_code if challan_for_annexure else ''
                _style_from_template(ws_annexure, r, 21, template_annexure_row).value = challan_for_annexure.challan_no if challan_for_annexure else ''
                _style_from_template(ws_annexure, r, 22, template_annexure_row).value = challan_text
                _style_from_template(ws_annexure, r, 23, template_annexure_row).value = None
                _style_from_template(ws_annexure, r, 24, template_annexure_row).value = ''
                _style_from_template(ws_annexure, r, 25, template_annexure_row).value = 'Book Entry' if (challan_for_annexure and challan_for_annexure.book_entry == 'yes') else ''
                _style_from_template(ws_annexure, r, 26, template_annexure_row).value = ''
                emp_code = getattr(emp, 'employee_code', False) or getattr(emp, 'barcode', False) or getattr(emp, 'identification_id', False) or ''
                _style_from_template(ws_annexure, r, 27, template_annexure_row).value = emp_code

                row_idx += 1
                deductee_sr += 1

        output = io.BytesIO()
        try:
            wb.save(output)
        except Exception as e:
            raise UserError(f"Failed to generate XLSX: {e}")

        output.seek(0)
        data = output.getvalue()
        output.close()

        filename = f"TDS_Return_{fields.Date.to_string(self.date_from)}_to_{fields.Date.to_string(self.date_to)}.xlsx"
        self.write({
            'output_file': base64.b64encode(data),
            'output_file_name': filename,
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_download(self):
        self.ensure_one()
        if not self.output_file:
            raise UserError('No file available. Please Generate first.')
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model={self._name}&id={self.id}&field=output_file&download=true&filename={self.output_file_name}',
            'target': 'self',
        }

    def action_new(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }
